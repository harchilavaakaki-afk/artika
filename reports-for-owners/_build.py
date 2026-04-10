"""Генерация XLSX и CSV на основе шаблона с новой колонкой 03-09.04.26."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import csv, pathlib

BASE = pathlib.Path(r"C:\Users\harch\OneDrive\Рабочий стол\cloude\Artika\reports-for-owners")

# Даты периодов (20 существующих + 1 новая)
LEAD_PERIODS = [
    "31.10.25-6.11.25","7-13.11.25","14-20.11.25","21-27.11.25","28.11-4.12.25",
    "5-11.12.25","12-18.12.25","19-25.12.25","26.12-15.01.26","16-22.01.2026",
    "23-29.01.2026","30.01-5.02.2026","6-12.02.2026","13-19.02.2026","20-26.02.2026",
    "27.02-5.03.2026","12.03.2026","13-19.03.26","20-26.03.26","27.03-2.04.26","03-09.04.26",
]
LEAD_VALUES = [35,56,60,57,46,48,54,45,155,62,70,55,56,62,42,29,41,39,56,53, 53]

# ТГ периоды (19 существующих + 1 новая). У ТГ первая дата 7-13.11.25 (без 31.10)
TG_PERIODS = LEAD_PERIODS[1:]  # начиная с 7-13.11.25
TG_VALUES  = [89,123,67,78,92,77,111,1185,118,132,121,114,101,86,97,60,91,60,63, 84]

# Подписчики — 20 дат + 1 новая
SUB_DATES  = [
    "на 10.11.25","11/13/2025","11/20/2025","11/27/2025","12/4/2025","11.12.2025",
    "18.12.2025","25.12.2025","15.01.2026","22.01.2026","29.01.2026","05.02.2026",
    "12.02.2026","19.02.2026","26.02.2026","05.03.2026","12.03.2026","19.03.2026",
    "26.03.2026","02.04.2026","09.04.2026",
]
SUB_VALUES = [229,390,514,534,565,582,580,601,711,726,768,786,811,827,831,844,850,851,867,870, 875]

NEW_COL_FILL = PatternFill(start_color="FFF6D6", end_color="FFF6D6", fill_type="solid")
NEW_HDR_FILL = PatternFill(start_color="FFEEA1", end_color="FFEEA1", fill_type="solid")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()
ws = wb.active
ws.title = "Отчёт"

# Шапка (как в оригинале, row 2-3)
ws["B2"] = "сайт padelvidnoe.ru"
ws["B2"].font = Font(bold=True, size=13)
ws["B3"] = "статистика по просмотрам 03.04-09.04.2026"
ws["B3"].font = Font(size=11)

# Лиды — row 24 header, row 25 values
LEAD_HDR_ROW = 24
LEAD_VAL_ROW = 25
ws.cell(row=LEAD_HDR_ROW, column=2, value="лидогенерация: количество заявок и звонков, поступивших через сайт").alignment = Alignment(wrap_text=True, vertical="center")
ws.cell(row=LEAD_HDR_ROW, column=2).font = BOLD
for i, (p, v) in enumerate(zip(LEAD_PERIODS, LEAD_VALUES)):
    col = 3 + i  # C..W
    c_hdr = ws.cell(row=LEAD_HDR_ROW, column=col, value=p)
    c_val = ws.cell(row=LEAD_VAL_ROW, column=col, value=v)
    c_hdr.alignment = CENTER; c_val.alignment = CENTER
    c_hdr.border = BORDER; c_val.border = BORDER
    c_hdr.font = Font(size=10)
    c_val.font = Font(bold=True, size=11)
    if i == len(LEAD_PERIODS) - 1:  # новая колонка
        c_hdr.fill = NEW_HDR_FILL
        c_val.fill = NEW_COL_FILL

# ТГ — row 27 header, row 28 values (первая дата в колонке D т.к. в шаблоне C пуста)
TG_HDR_ROW = 27
TG_VAL_ROW = 28
ws.cell(row=TG_HDR_ROW, column=2, value="переходы в ТГ канал").font = BOLD
for i, (p, v) in enumerate(zip(TG_PERIODS, TG_VALUES)):
    col = 4 + i  # D..W
    c_hdr = ws.cell(row=TG_HDR_ROW, column=col, value=p)
    c_val = ws.cell(row=TG_VAL_ROW, column=col, value=v)
    c_hdr.alignment = CENTER; c_val.alignment = CENTER
    c_hdr.border = BORDER; c_val.border = BORDER
    c_hdr.font = Font(size=10)
    c_val.font = Font(bold=True, size=11)
    if i == len(TG_PERIODS) - 1:
        c_hdr.fill = NEW_HDR_FILL
        c_val.fill = NEW_COL_FILL

# Подписчики — row 33 link, row 34 dates header, row 35 values
ws["B33"] = "https://t.me/padelvidnoe"
ws["B33"].font = Font(color="0563C1", underline="single")

SUB_HDR_ROW = 34
SUB_VAL_ROW = 35
ws.cell(row=SUB_HDR_ROW, column=2, value="количество подписчиков").font = BOLD
for i, (d, v) in enumerate(zip(SUB_DATES, SUB_VALUES)):
    col = 3 + i  # C..W
    c_hdr = ws.cell(row=SUB_HDR_ROW, column=col, value=d)
    c_val = ws.cell(row=SUB_VAL_ROW, column=col, value=v)
    c_hdr.alignment = CENTER; c_val.alignment = CENTER
    c_hdr.border = BORDER; c_val.border = BORDER
    c_hdr.font = Font(size=10)
    c_val.font = Font(bold=True, size=11)
    if i == len(SUB_DATES) - 1:
        c_hdr.fill = NEW_HDR_FILL
        c_val.fill = NEW_COL_FILL

# что делаем
ws["B39"] = "что делаем:"
ws["B39"].font = BOLD
ws["B40"] = "проработка контента и рубрик с правилами, тренерами, живые видео с игроками"
ws["B41"] = "подготовка контент-плана"
ws["B42"] = "02.04 — запущена акция «Приведи друга — 1 000 ₽ на депозит»"
ws["B43"] = "03.04 — анонсированы технические тренировки с тренером Корболиным С. (11/18/25 апреля)"
ws["B44"] = "06.04 — опубликованы итоги Кубка ПРОЛИНК"
ws["B45"] = "07.04 — анонсированы турниры: Rainbow 11.04, Пятничный Americano 17.04, Кубок Seven Six 18.04"

# Ширина колонок
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 35
for col in range(3, 24):  # C..W
    ws.column_dimensions[get_column_letter(col)].width = 12

# Высота строк-шапок
ws.row_dimensions[LEAD_HDR_ROW].height = 40
ws.row_dimensions[TG_HDR_ROW].height = 24
ws.row_dimensions[SUB_HDR_ROW].height = 24

out_xlsx = BASE / "report_03-09-04-2026.xlsx"
wb.save(out_xlsx)
print("XLSX saved:", out_xlsx)

# CSV версия (только данные + новая колонка)
out_csv = BASE / "report_03-09-04-2026.csv"
with open(out_csv, "w", encoding="utf-8-sig", newline="") as f:
    w = csv.writer(f)
    for _ in range(1): w.writerow([])
    w.writerow(["", "сайт padelvidnoe.ru"])
    w.writerow(["", "статистика по просмотрам 03.04-09.04.2026"])
    for _ in range(20): w.writerow([])
    # Лиды
    w.writerow(["", "лидогенерация: количество заявок и звонков, поступивших через сайт"] + LEAD_PERIODS)
    w.writerow(["", ""] + LEAD_VALUES)
    w.writerow([])
    # ТГ — в колонке C пусто, даты начинаются с D
    w.writerow(["", "переходы в ТГ канал", ""] + TG_PERIODS)
    w.writerow(["", "", ""] + TG_VALUES)
    for _ in range(4): w.writerow([])
    w.writerow(["", "https://t.me/padelvidnoe"])
    w.writerow(["", "количество подписчиков"] + SUB_DATES)
    w.writerow(["", ""] + SUB_VALUES)
    for _ in range(3): w.writerow([])
    w.writerow(["", "что делаем:"])
    w.writerow(["", "проработка контента и рубрик с правилами, тренерами, живые видео с игроками"])
    w.writerow(["", "подготовка контент-плана"])
    w.writerow(["", "02.04 — запущена акция «Приведи друга — 1 000 ₽ на депозит»"])
    w.writerow(["", "03.04 — анонсированы технические тренировки с тренером Корболиным С."])
    w.writerow(["", "06.04 — опубликованы итоги Кубка ПРОЛИНК"])
    w.writerow(["", "07.04 — анонсированы турниры: Rainbow 11.04, Americano 17.04, Seven Six 18.04"])
print("CSV saved:", out_csv)

print()
print("=== Координаты для ручного ввода в Google Sheets ===")
print(" B3  -> статистика по просмотрам 03.04-09.04.2026")
print(" W24 -> 03-09.04.26")
print(" W25 -> 53")
print(" W27 -> 03-09.04.26")
print(" W28 -> 84")
print(" W34 -> 09.04.2026")
print(" W35 -> 875")
