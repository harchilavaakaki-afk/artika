#!/usr/bin/env python3
"""
Генератор 4-вкладочного XLSX-отчёта для собственника.
Читает: _live_data.json + _manual_input.json + _history.json
Пишет: report_DD-MM-YYYY.xlsx

Использование:
  python generate_report.py 2026-04-10 2026-04-16
"""
import sys, json, pathlib, datetime, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

BASE = pathlib.Path(__file__).parent

# ─── Загрузка данных ──────────────────────────────────────────────────────────
def load(path):
    p = BASE / path
    return json.loads(p.read_text("utf-8")) if p.exists() else {}

# ─── Стили ───────────────────────────────────────────────────────────────────
THIN_SIDE  = Side(border_style="thin", color="CCCCCC")
BORDER     = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

FILL_NEW_HDR  = fill("FFEEA1")
FILL_NEW_VAL  = fill("FFF6D6")
FILL_SECTION  = fill("F5F4F2")
FILL_HEADER   = fill("E8F4FD")
FILL_DIRECT   = fill("FFF3E0")
FILL_EVENTS   = fill("F3E5F5")

def hdr_font(size=11, bold=True, color="2B2A28"):
    return Font(bold=bold, size=size, color=color)

def write_cell(ws, row, col, value, **kw):
    c = ws.cell(row=row, column=col, value=value)
    if "font"   in kw: c.font      = kw["font"]
    if "fill"   in kw: c.fill      = kw["fill"]
    if "align"  in kw: c.alignment = kw["align"]
    if "border" in kw: c.border    = kw["border"]
    return c

def apply_series_row(ws, row, label_col, data_cols, values, hdr_fill=None, val_fill=None):
    """Рисует строку с label + данными по колонкам."""
    for i, (col, val) in enumerate(zip(data_cols, values)):
        is_new = (i == len(values) - 1)
        hf = FILL_NEW_HDR if is_new else (hdr_fill or fill("F5F4F2"))
        vf = FILL_NEW_VAL if is_new else (val_fill or fill("FFFFFF"))
        c = ws.cell(row=row, column=col, value=val)
        c.fill = vf if row % 2 == 0 else vf
        c.border = BORDER
        c.alignment = CENTER
        c.font = Font(bold=is_new, size=11 if is_new else 10)
        if is_new:
            c.fill = FILL_NEW_VAL

# ─── Вкладка 1: Сайт и ТГ ────────────────────────────────────────────────────
def build_tab1(wb, live, manual, history, date1, date2):
    ws = wb.active
    ws.title = "Сайт и ТГ"
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38

    d1 = datetime.date.fromisoformat(date1)
    d2 = datetime.date.fromisoformat(date2)
    label    = manual.get("period_label", f"{d1.strftime('%d.%m')}-{d2.strftime('%d.%m.%y')}")
    hdr_text = f"статистика по просмотрам {d1.strftime('%d.%m')}-{d2.strftime('%d.%m.%Y')}"

    # Заголовок
    ws["B2"] = "сайт padelvidnoe.ru"
    ws["B2"].font = Font(bold=True, size=13)
    ws["B3"] = hdr_text
    ws["B3"].font = Font(size=11, color="6B6B6B")

    # Данные из live + manual
    m  = live.get("metrika", {})
    tg = live.get("tg", {})
    ov = manual.get("tab1_overrides", {})

    leads       = ov.get("leads") or m.get("leads", 0)
    calls       = ov.get("calls") or m.get("calls", 0)
    tg_clicks   = ov.get("tg_clicks") or m.get("tg_clicks", 0)
    subscribers = ov.get("subscribers") or tg.get("subscribers", 0)

    # Серии: добавляем новый период
    lead_periods = history["leads"]["periods"] + [label]
    lead_values  = history["leads"]["values"]  + [leads + calls]
    tg_periods   = history["tg_clicks"]["periods"] + [label]
    tg_values    = history["tg_clicks"]["values"]  + [tg_clicks]
    sub_dates    = history["subscribers"]["dates"]  + [manual.get("period_end_date", d2.strftime("%d.%m.%Y"))]
    sub_values   = history["subscribers"]["values"] + [subscribers]

    # Лиды — строки 24-25
    ROW_LEAD_HDR, ROW_LEAD_VAL = 24, 25
    ws.cell(ROW_LEAD_HDR, 2, "лидогенерация: количество заявок и звонков, поступивших через сайт").font = hdr_font()
    ws.row_dimensions[ROW_LEAD_HDR].height = 40
    data_cols = list(range(3, 3 + len(lead_periods)))
    for i, (col, p, v) in enumerate(zip(data_cols, lead_periods, lead_values)):
        is_new = (i == len(lead_periods) - 1)
        c_h = ws.cell(ROW_LEAD_HDR, col, p)
        c_v = ws.cell(ROW_LEAD_VAL, col, v)
        for c in (c_h, c_v):
            c.border = BORDER; c.alignment = CENTER
        c_h.font = Font(size=9, bold=is_new)
        c_v.font = Font(size=12, bold=True)
        if is_new:
            c_h.fill = FILL_NEW_HDR
            c_v.fill = FILL_NEW_VAL
    for col in data_cols:
        ws.column_dimensions[get_column_letter(col)].width = 13

    # ТГ переходы — строки 27-28
    ROW_TG_HDR, ROW_TG_VAL = 27, 28
    ws.cell(ROW_TG_HDR, 2, "переходы в ТГ канал").font = hdr_font()
    tg_cols = list(range(4, 4 + len(tg_periods)))  # начинается с D
    for i, (col, p, v) in enumerate(zip(tg_cols, tg_periods, tg_values)):
        is_new = (i == len(tg_periods) - 1)
        c_h = ws.cell(ROW_TG_HDR, col, p)
        c_v = ws.cell(ROW_TG_VAL, col, v)
        for c in (c_h, c_v):
            c.border = BORDER; c.alignment = CENTER
        c_h.font = Font(size=9, bold=is_new)
        c_v.font = Font(size=12, bold=True)
        if is_new:
            c_h.fill = FILL_NEW_HDR
            c_v.fill = FILL_NEW_VAL

    # Метрика источники — строки 6-14
    ws["B6"] = "Источник трафика"
    ws["B6"].font = hdr_font()
    ws["B6"].fill = FILL_HEADER
    for col, hdr in [(3,"Визиты"),(4,"Польз."),(5,"Отказы %"),(6,"Глубина"),(7,"Время сек.")]:
        c = ws.cell(6, col, hdr)
        c.font = Font(bold=True, size=10); c.fill = FILL_HEADER; c.alignment = CENTER
    sources = m.get("sources", [])
    for i, s in enumerate(sources):
        r = 7 + i
        ws.cell(r, 2, s["name"]).font = Font(size=11)
        for col, key in [(3,"visits"),(4,"users"),(5,"bounce"),(6,"depth"),(7,"duration")]:
            c = ws.cell(r, col, s.get(key))
            c.alignment = CENTER; c.font = Font(size=11)
        ws.row_dimensions[r].height = 18
    if not sources:
        ws["C7"] = "← данные получи через connector_metrika.py"
        ws["C7"].font = Font(italic=True, color="999999")

    # Подписчики — строки 33-35
    ws["B33"] = "https://t.me/padelvidnoe"
    ws["B33"].font = Font(color="0563C1", underline="single", size=11)
    ws.cell(34, 2, "количество подписчиков").font = hdr_font()
    sub_cols = list(range(3, 3 + len(sub_dates)))
    for i, (col, d, v) in enumerate(zip(sub_cols, sub_dates, sub_values)):
        is_new = (i == len(sub_dates) - 1)
        c_h = ws.cell(34, col, d)
        c_v = ws.cell(35, col, v)
        for c in (c_h, c_v):
            c.border = BORDER; c.alignment = CENTER
        c_h.font = Font(size=9, bold=is_new)
        c_v.font = Font(size=12, bold=True)
        if is_new:
            c_h.fill = FILL_NEW_HDR
            c_v.fill = FILL_NEW_VAL

    # Прирост подписчиков
    growth = tg.get("growth")
    if growth is not None:
        sign = "+" if growth >= 0 else ""
        ws.cell(36, 2, f"Прирост за неделю: {sign}{growth}").font = Font(italic=True, size=10, color="555555")

    # Что делаем
    ws.cell(39, 2, "что делаем:").font = hdr_font(size=12)
    for i, line in enumerate(manual.get("tab1_what_we_did", [])):
        c = ws.cell(40 + i, 2, line)
        c.font = Font(size=11); c.alignment = LEFT_WRAP
        ws.row_dimensions[40 + i].height = 22


# ─── Вкладка 2: Директ ───────────────────────────────────────────────────────
def build_tab2(wb, live, manual, date1, date2):
    ws = wb.create_sheet("Директ")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 16  # Дата
    for col_i, w in enumerate([10,8,8,14,14,10,12,18], start=3):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    D_HDRS = ["Дата","Показы","Клики","CTR (%)","Расход (руб.)","Ср. цена клика","Глубина (стр.)","Конверсия %"]
    D_KEYS = ["date","impressions","clicks","ctr","cost","avg_cpc","depth","conv_rate"]

    campaigns = live.get("direct", [])
    cur_row = 2

    for camp in campaigns:
        cid  = camp.get("id", "")
        name = camp.get("name", "Кампания")
        tot  = camp.get("totals", {})
        daily= camp.get("daily", [])

        d1 = datetime.date.fromisoformat(date1).strftime("%d.%m.%Y")
        d2 = datetime.date.fromisoformat(date2).strftime("%d.%m.%Y")

        # Заголовок кампании
        title = f"Кампания «{name}» ({cid}), период {d1} - {d2}"
        c = ws.cell(cur_row, 2, title)
        c.font = Font(bold=True, size=12); c.fill = FILL_DIRECT
        cur_row += 1

        # Итого
        ws.cell(cur_row, 2, "Всего:").font = Font(bold=True, size=11)
        tot_vals = [
            f"{d1} по {d2}",
            tot.get("impressions"),tot.get("clicks"),tot.get("ctr"),
            tot.get("cost"),tot.get("avg_cpc"),tot.get("depth"),tot.get("conv_rate")
        ]
        for col_i, v in enumerate(tot_vals, start=2):
            c = ws.cell(cur_row, col_i, v)
            c.font = Font(bold=True, size=10); c.fill = fill("FFEDB3"); c.alignment = CENTER; c.border = BORDER
        cur_row += 1

        # Шапка таблицы
        for col_i, h in enumerate(D_HDRS, start=2):
            c = ws.cell(cur_row, col_i, h)
            c.font = Font(bold=True, size=10); c.fill = FILL_SECTION; c.alignment = CENTER; c.border = BORDER
        cur_row += 1

        # Данные по дням
        if daily:
            for row_data in daily:
                for col_i, key in enumerate(D_KEYS, start=2):
                    c = ws.cell(cur_row, col_i, row_data.get(key))
                    c.alignment = CENTER; c.border = BORDER; c.font = Font(size=10)
                cur_row += 1
        else:
            ws.cell(cur_row, 2, "← данные получи через connector_direct.py").font = Font(italic=True, color="999999")
            cur_row += 1

        cur_row += 2  # разрыв между кампаниями

    # Что делаем
    ws.cell(cur_row, 2, "Что делаем:").font = hdr_font(size=12)
    cur_row += 1
    for line in manual.get("tab2_what_we_did", []):
        c = ws.cell(cur_row, 2, line)
        c.font = Font(size=11); c.alignment = LEFT_WRAP; ws.row_dimensions[cur_row].height = 22
        cur_row += 1


# ─── Вкладка 3: Реклама и продвижение соц сети ───────────────────────────────
def build_tab3(wb, manual, date1, date2):
    ws = wb.create_sheet("Реклама и соц сети")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 70

    d1 = datetime.date.fromisoformat(date1).strftime("%d.%m")
    d2 = datetime.date.fromisoformat(date2).strftime("%d.%m.%y")
    label = manual.get("period_label", f"{d1}-{d2}")

    # Заголовок
    c = ws.cell(2, 2, f"Реклама и продвижение в соц. сетях — {label}")
    c.font = Font(bold=True, size=13)

    ws.cell(3, 2, "мероприятия и движения по сайту, телеграмм, промовидео, акции в сети, соц. сети").font = Font(size=10, italic=True, color="6B6B6B")

    # Шапка таблицы
    for col_i, h in [(2, "Направление"), (3, "Описание / Результат")]:
        c = ws.cell(5, col_i, h)
        c.font = Font(bold=True, size=11); c.fill = FILL_SECTION; c.alignment = CENTER; c.border = BORDER

    promo_items = manual.get("tab3_promo", [])
    cur_row = 6
    for item in promo_items:
        c_b = ws.cell(cur_row, 2, item.get("title", ""))
        c_b.font = Font(bold=True, size=11); c_b.alignment = LEFT_WRAP; c_b.border = BORDER

        c_d = ws.cell(cur_row, 3, item.get("details", ""))
        c_d.font = Font(size=11); c_d.alignment = LEFT_WRAP; c_d.border = BORDER

        ws.row_dimensions[cur_row].height = 45
        cur_row += 1

    if not promo_items:
        ws.cell(6, 2, "← заполни tab3_promo в _manual_input.json").font = Font(italic=True, color="999999")


# ─── Вкладка 4: Продвижение / мероприятия ────────────────────────────────────
def build_tab4(wb, manual, date1, date2):
    ws = wb.create_sheet("Мероприятия")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 18

    d1 = datetime.date.fromisoformat(date1).strftime("%d.%m")
    d2 = datetime.date.fromisoformat(date2).strftime("%d.%m.%y")
    label = manual.get("period_label", f"{d1}-{d2}")

    c = ws.cell(2, 2, f"Продвижение / мероприятия — {label}")
    c.font = Font(bold=True, size=13)

    ws.cell(3, 2, "конкретные мероприятия и акции, проведённые в падел-центре").font = Font(size=10, italic=True, color="6B6B6B")

    # Шапка
    hdrs = ["Мероприятие / Акция", "Дата", "Описание", "Участники"]
    for col_i, h in enumerate(hdrs, start=2):
        c = ws.cell(5, col_i, h)
        c.font = Font(bold=True, size=11); c.fill = FILL_EVENTS; c.alignment = CENTER; c.border = BORDER

    events = manual.get("tab4_events", [])
    cur_row = 6
    for ev in events:
        vals = [ev.get("title",""), ev.get("date",""), ev.get("details",""), ev.get("attendance","")]
        for col_i, v in enumerate(vals, start=2):
            c = ws.cell(cur_row, col_i, v)
            c.font = Font(size=11); c.alignment = LEFT_WRAP; c.border = BORDER
        ws.row_dimensions[cur_row].height = 50
        cur_row += 1

    if not events:
        ws.cell(6, 2, "← заполни tab4_events в _manual_input.json").font = Font(italic=True, color="999999")

    ws.cell(cur_row + 1, 2, "* данные об участниках уточнять у @olesia_1908").font = Font(size=9, italic=True, color="999999")


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 3:
        print("Использование: python generate_report.py 2026-04-10 2026-04-16")
        sys.exit(1)

    date1, date2 = sys.argv[1], sys.argv[2]
    live    = load("_live_data.json")
    manual  = load("_manual_input.json")
    history = load("_history.json")

    d1 = datetime.date.fromisoformat(date1)
    d2 = datetime.date.fromisoformat(date2)

    wb = Workbook()
    build_tab1(wb, live, manual, history, date1, date2)
    build_tab2(wb, live, manual, date1, date2)
    build_tab3(wb, manual, date1, date2)
    build_tab4(wb, manual, date1, date2)

    fname = f"report_{d1.strftime('%d-%m')}-{d2.strftime('%d-%m-%Y')}.xlsx"
    out = BASE / fname
    wb.save(out)
    print(f"✓ Отчёт сохранён: {out}")

    # Обновляем _history.json новой неделей
    update_history(history, manual, live, date1, date2)

def update_history(history, manual, live, date1, date2):
    """Добавляет новую неделю в _history.json (если её ещё нет)."""
    label = manual.get("period_label", "")
    if not label or label in history["leads"]["periods"]:
        return  # уже добавлено

    m  = live.get("metrika", {})
    tg = live.get("tg", {})
    ov = manual.get("tab1_overrides", {})

    leads       = ov.get("leads") or m.get("leads", 0)
    calls       = ov.get("calls") or m.get("calls", 0)
    tg_clicks   = ov.get("tg_clicks") or m.get("tg_clicks", 0)
    subscribers = ov.get("subscribers") or tg.get("subscribers", 0)
    end_date    = manual.get("period_end_date", datetime.date.fromisoformat(date2).strftime("%d.%m.%Y"))

    history["leads"]["periods"].append(label)
    history["leads"]["values"].append(leads + calls)

    if label not in history["tg_clicks"]["periods"]:
        history["tg_clicks"]["periods"].append(label)
        history["tg_clicks"]["values"].append(tg_clicks)

    if end_date not in history["subscribers"]["dates"]:
        history["subscribers"]["dates"].append(end_date)
        history["subscribers"]["values"].append(subscribers)

    (BASE / "_history.json").write_text(json.dumps(history, ensure_ascii=False, indent=2), "utf-8")
    print("✓ _history.json обновлён новой неделей")

if __name__ == "__main__":
    main()
